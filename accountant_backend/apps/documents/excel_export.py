"""Data Export (Excel) — one sheet per business: business name up top,
then every customer's statement one after another, then grand totals at
the bottom. Built with pandas since the numbers/dates already come out of
the ORM as a natural table per customer; openpyxl (pandas' xlsx engine)
does the light styling (bold headers, column widths) after the data is
written.
"""

import pandas as pd
from io import BytesIO

from apps.customers.models import Customer
from apps.sales.models import ActivityEntry

COLUMNS = ["Date", "Type", "Details", "Amount", "Balance"]


def _entry_row(entry):
    if entry.type == "sale":
        items = ", ".join(f"{li.item_name} x{li.quantity}" for li in entry.line_items.all())
        details = items or (entry.note or "")
    else:
        details = entry.payment_method or entry.note or ""
    return {
        "Date": entry.timestamp.strftime("%Y-%m-%d"),
        "Type": "Sale" if entry.type == "sale" else "Payment",
        "Details": details,
        "Amount": float(entry.amount),
        "Balance": float(entry.balance_after),
    }


def build_export_workbook(business) -> bytes:
    customers = Customer.objects.filter(business=business).order_by("name")

    output = BytesIO()
    total_sale = 0
    total_payment = 0
    total_balance = 0

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = "Statement"
        pd.DataFrame().to_excel(writer, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        ws.delete_cols(1, 1)

        row = 1
        ws.cell(row=row, column=1, value=business.business_name)
        ws.cell(row=row, column=1).font = _bold(size=14)
        row += 2

        for customer in customers:
            entries = (
                ActivityEntry.objects.filter(business=business, customer=customer)
                .prefetch_related("line_items")
                .order_by("timestamp", "id")
            )
            rows = [_entry_row(e) for e in entries]
            for e in entries:
                if e.type == "sale":
                    total_sale += float(e.amount)
                else:
                    total_payment += float(e.amount)
            total_balance += float(customer.current_balance)

            ws.cell(row=row, column=1, value=f"Customer: {customer.name}")
            ws.cell(row=row, column=1).font = _bold(size=12)
            row += 1

            df = pd.DataFrame(rows, columns=COLUMNS)
            df.to_excel(writer, sheet_name=sheet_name, startrow=row - 1, index=False)
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row, column=col_idx).font = _bold()
            row += len(df) + 1

            ws.cell(row=row, column=1, value="Closing Balance")
            ws.cell(row=row, column=2, value=float(customer.current_balance))
            ws.cell(row=row, column=1).font = _bold()
            ws.cell(row=row, column=2).font = _bold()
            row += 2

        ws.cell(row=row, column=1, value="Total Sale")
        ws.cell(row=row, column=2, value=total_sale)
        row += 1
        ws.cell(row=row, column=1, value="Total Amount Received")
        ws.cell(row=row, column=2, value=total_payment)
        row += 1
        ws.cell(row=row, column=1, value="Total Balance (All Customers)")
        ws.cell(row=row, column=2, value=total_balance)
        for r in range(row - 2, row + 1):
            ws.cell(row=r, column=1).font = _bold()
            ws.cell(row=r, column=2).font = _bold()

        for col_idx, width in zip(range(1, 6), (14, 10, 30, 12, 12)):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    output.seek(0)
    return output.getvalue()


def _bold(size=11):
    from openpyxl.styles import Font

    return Font(bold=True, size=size)
